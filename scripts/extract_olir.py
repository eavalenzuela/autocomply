#!/usr/bin/env python3
"""Extract the NIST OLIR 800-53r5 -> ISO/IEC 27001:2022 mapping into a flat TSV.

Source: the public-domain NIST OLIR submission spreadsheet (referenceId 155 in
the OLIR catalog), vendored under data/vendor/olir/. One sheet per 800-53 family;
each row pairs a focal 800-53 element with a reference ISO 27001:2022 element
(Annex A control "A.5.x" or management clause "5.2"). The submission is set-based
(it lists relatedness; per-row relationship type / strength are not populated).

Output: data/vendor/olir/olir-800-53r5-iso27001-2022.tsv  (control<TAB>iso_ref),
normalized to our canonical codes (AC-02(01) -> AC-2(1)) and stdlib-readable by
gen_crosswalk.py, so the crosswalk generator needs no xlsx dependency.

Requires openpyxl (pip install openpyxl). Re-run only when the vendored xlsx is
refreshed.
"""
import re
import pathlib
import openpyxl

ROOT = pathlib.Path(__file__).resolve().parent.parent
XLSX = ROOT / "data/vendor/olir/sp800-53r5-to-iso27001-2022-olir.xlsx"
OUT = ROOT / "data/vendor/olir/olir-800-53r5-iso27001-2022.tsv"

# Our ISO YAML splits clause 9.2 into 9.2.1/9.2.2; OLIR cites the parent "9.2".
ISO_REF_FIXUP = {"9.2": "9.2.1"}


def norm_control(code: str):
    """AC-01 -> AC-1 ; AC-02(01) -> AC-2(1). Returns None if unparseable."""
    m = re.match(r"^([A-Z]{2})-(\d+)(?:\((\d+)\))?$", str(code).strip())
    if not m:
        return None
    fam, num, enh = m.groups()
    out = f"{fam}-{int(num)}"
    return out + (f"({int(enh)})" if enh is not None else "")


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    pairs = set()
    skipped_focal = set()
    for sheet in wb.sheetnames:
        if sheet == "Definitions":
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            focal = row[0]
            ref = row[3] if len(row) > 3 else None
            if focal is None or ref is None:
                continue
            control = norm_control(focal)
            if control is None:
                skipped_focal.add(str(focal))
                continue
            iso = str(ref).strip()
            iso = ISO_REF_FIXUP.get(iso, iso)
            pairs.add((control, iso))

    rows = sorted(pairs, key=lambda p: (p[0], p[1]))
    body = "# control\tiso_ref  (from NIST OLIR 800-53r5 -> ISO 27001:2022, public domain)\n"
    body += "\n".join(f"{c}\t{i}" for c, i in rows) + "\n"
    OUT.write_text(body)
    controls = {c for c, _ in rows}
    print(f"wrote {OUT.relative_to(ROOT)}: {len(rows)} pairs, {len(controls)} controls")
    if skipped_focal:
        print(f"skipped unparseable focal codes: {sorted(skipped_focal)}")


if __name__ == "__main__":
    main()
